"""Generate Excel report with 4 NLDM tables per cell."""

import json
import glob
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results", "nldm")
OUTPUT_FILE = os.path.join(RESULTS_DIR, "nldm_tables.xlsx")

TABLE_NAMES = ["cell_rise", "cell_fall", "rise_transition", "fall_transition"]
TABLE_LABELS = {
    "cell_rise":        "Cell Rise Delay (ns)",
    "cell_fall":        "Cell Fall Delay (ns)",
    "rise_transition":  "Rise Transition Time (ns)",
    "fall_transition":  "Fall Transition Time (ns)",
}

# Styles
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
TITLE_FONT   = Font(bold=True, size=11, color="FFFFFF")
TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")
INDEX_FILL   = PatternFill("solid", fgColor="D6E4F0")
INDEX_FONT   = Font(bold=True, size=9)
DATA_FONT    = Font(size=9)
CENTER       = Alignment(horizontal="center", vertical="center")
THIN         = Side(style="thin", color="AAAAAA")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_table(ws, data, row_labels, col_labels, top_row, left_col, title):
    """Write one 7x7 NLDM table at (top_row, left_col)."""
    n_rows = len(row_labels)
    n_cols = len(col_labels)

    # Title row spanning whole table (+1 for the index column)
    ws.merge_cells(
        start_row=top_row, start_column=left_col,
        end_row=top_row,   end_column=left_col + n_cols
    )
    title_cell = ws.cell(row=top_row, column=left_col, value=title)
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER
    title_cell.border    = THIN_BORDER

    # Corner cell header
    corner = ws.cell(row=top_row + 1, column=left_col,
                     value="tr \\ Cload")
    corner.font      = HEADER_FONT
    corner.fill      = HEADER_FILL
    corner.alignment = CENTER
    corner.border    = THIN_BORDER

    # Column headers (load cap values)
    for j, cap in enumerate(col_labels):
        c = ws.cell(row=top_row + 1, column=left_col + 1 + j,
                    value=round(cap, 4))
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = THIN_BORDER

    # Row headers + data
    for i, (tr, row_data) in enumerate(zip(row_labels, data)):
        # Row index
        idx = ws.cell(row=top_row + 2 + i, column=left_col,
                      value=round(tr, 4))
        idx.font      = INDEX_FONT
        idx.fill      = INDEX_FILL
        idx.alignment = CENTER
        idx.border    = THIN_BORDER

        for j, val in enumerate(row_data):
            d = ws.cell(row=top_row + 2 + i, column=left_col + 1 + j,
                        value=round(val, 6))
            d.font      = DATA_FONT
            d.alignment = CENTER
            d.border    = THIN_BORDER

    # Column widths
    ws.column_dimensions[get_column_letter(left_col)].width = 10
    for j in range(n_cols):
        ws.column_dimensions[get_column_letter(left_col + 1 + j)].width = 10


def process_cell(wb, json_path):
    with open(json_path) as f:
        data = json.load(f)

    cell_name        = data["cell"]
    input_transitions = data["input_transition_ns"]
    load_caps        = data["load_cap_pf"]
    tables           = data["tables_ns"]

    ws = wb.create_sheet(title=cell_name)
    ws.sheet_view.showGridLines = True

    # Place 2 tables per row, with a gap between them
    # Layout: table1 at col 1, table2 at col 10, second pair starts at row 12
    positions = [
        (1,  1),   # cell_rise
        (1,  11),  # cell_fall
        (12, 1),   # rise_transition
        (12, 11),  # fall_transition
    ]

    # Cell name banner at very top
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    banner = ws.cell(row=1, column=1, value=f"NLDM Tables — {cell_name.upper()}")
    banner.font      = Font(bold=True, size=13, color="FFFFFF")
    banner.fill      = PatternFill("solid", fgColor="1A3A5C")
    banner.alignment = CENTER

    row_offset = 2  # shift tables down one row for the banner

    for idx, tname in enumerate(TABLE_NAMES):
        base_row, base_col = positions[idx]
        write_table(
            ws,
            tables[tname],
            input_transitions,
            load_caps,
            top_row=base_row + row_offset,
            left_col=base_col,
            title=TABLE_LABELS[tname],
        )

    # Row heights
    for r in range(1, 25):
        ws.row_dimensions[r].height = 18

    # Axis labels (units reminder)
    note_row = 23
    ws.cell(row=note_row, column=1,
            value="Rows: Input Transition (ns)    |    Columns: Load Capacitance (pF)")
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="555555")


def main():
    json_files = sorted(glob.glob(os.path.join(RESULTS_DIR, "*.json")))
    if not json_files:
        print("No JSON files found in", RESULTS_DIR)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for jf in json_files:
        print(f"Processing {os.path.basename(jf)} ...")
        process_cell(wb, jf)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
